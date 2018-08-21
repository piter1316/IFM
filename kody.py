from openpyxl import load_workbook

print('Ładowanie listy produktów... ')
wb = load_workbook('IFM Preisliste 2018.xlsx')
print(wb.sheetnames)
sheet = wb['Preisliste 2018']

print('Zakończono. Porbrano: ', sheet.max_row, 'kodów')

kody = []
for i in range(9, sheet.max_row):
    kody.append(sheet.cell(row=i, column=1).value)
    # print(sheet.cell(row=i, column=1).value)
